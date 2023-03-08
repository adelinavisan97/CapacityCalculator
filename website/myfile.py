from openpyxl import load_workbook

workbook = load_workbook(filename="spreadsheet.xlsx")
workbook.sheetnames

sheet = workbook.active
one = []
for value in sheet.iter_rows(min_row=1,
                             max_row=4,
                             min_col=1,
                             max_col=3,
                             values_only=True):
    one.append(value)
print(one)
rowOne = one[0]
rowTwo = one[1]
rowThree = one[2]
rowFour = one[3]
print(rowOne)
colOne, colTwo, colThree = rowOne
rowOne = []
rowOne.append(colOne)
rowOne.append(colTwo)
rowOne.append(colThree)
print(rowOne)
sheet.title
